import toml
import openpyxl

# translate config to dictionary
configDict = toml.load('config.toml')

# geeting the worksheet to transform
sourceWorkbook = 
  openpyxl.load_workbook('./Spreadsheets' + configDict.get('filename'), read_only=True)

sourceConfig = configDict.get('sheet')
sourceRow = 0
sourceCol = 0
sourceSheet = 'Sheet 1' # default Sheet name in Excel

for key, value in sheet.items():
  if (str(key) == 'rows'):
    sourceRow = value
  elif (str(key) == 'cols'):
    sourceCol = value
  elif (str(key) == 'name'):
    sourceSheet = sourceWorkbook.get_sheet_by_name(str(value))


separator = configDict.get('separator')

# output excel destination
dest = "output_book.xlsx"
wb = openpyxl.Workbook()
outputSheet = wb.create_sheet(title="output_sheet")

headers = configDict.get('header')

# placing all the headers
for key, value in headers.items():
  columnIndex = openpyxl.utils.column_index_from_string(str(key))
  outputSheet.cell(row=1, column=columnIndex, value=str(value))

# iterating through the source sheet
for iterRow in range(1, sourceRow + 1):

  outputDict = {} # empty dictionary for every row
  for iterCol in range(1, sourceCol + 1):

    content = sourceSheet.cell(row = iterRow, column = iterCol)
    colLetter = openpyxl.utils.get_column_letter(iterCol)
        
    # check if the present column maps to something in the config file
    # default returnvalue is NoMap
    mapsTo = configDict.get(colLetter, 'NoMap')
        
    # If source doesn't map to anything, ignore it and move on
    if (mapsTo == 'NoMap'):
      continue
    else:

      # if the output column doesn't yet exist in the output hash table
      # create it
      # otherwise append to the existing array at for the map
      if (outputDict.get(mapsTo, 'None') == None):
        # creating a new array in case nothing exists
        outputDict[mapsTo] = [content]
      else:
        # appending to original array otherwise
        outputDict[mapsTo].append(content)

  for key, dictValue in outputDict.items():
    normalizedRow = iterRow + 2
    columnIndex = openpyxl.utils.column_index_from_string(key)

    # converting the contents of the cell into a string
    # separated by the specified separator
    formattedValue = separator.join(map(str, dictValue))
    # inserting cells into the new
    outputSheet.cell(row = normalizedRow, column = columnIndex, value = dictValue)




# TODO: transform the dictionary to a spreadsheet

# All the values in the cols are in sequential ascending order
# When transforming them to spreadsheets, just pasting them in 
# successive rows
